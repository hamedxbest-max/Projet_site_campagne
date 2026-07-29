from django.conf import settings
from django.db import models
from django.http import HttpResponse
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.response import Response
from rest_framework.views import APIView

from .taramoney import TaramoneyClient, TaramoneyError
import openpyxl
from openpyxl.styles import Font, PatternFill

from .models import Contribution, PaymentInstallment, MIN_INSTALLMENT, Payment
from .serializers import (
    ContributionSerializer,
    StudentCreateSerializer,
    DonorCreateSerializer,
    PaymentRequestSerializer,
    RegisteredStudentSerializer,
    STUDENT_FEE,
)
from . import campay


class ContributionViewSet(viewsets.ModelViewSet):
    """
    Public endpoints used by the React frontend:
      POST   /api/contributions/student/           -> inscription étudiant
      POST   /api/contributions/donor/               -> inscription donateur
      POST   /api/contributions/{id}/pay/            -> trigger Campay payment
      GET    /api/contributions/{id}/status/         -> poll payment status
      POST   /api/contributions/{id}/upload-proof/ -> preuve paiement espèces
      POST   /api/contributions/{id}/confirm-cash/ -> confirmer paiement espèces (sans Campay)

    Admin-only:
      GET    /api/contributions/                   -> list all (requires auth)
      GET    /api/contributions/export_excel/      -> download .xlsx (requires auth)
    """
    queryset = Contribution.objects.all()
    serializer_class = ContributionSerializer
    http_method_names = ['get', 'post', 'patch', 'head', 'options']

    def get_permissions(self):
        admin_actions = ('list', 'export_excel', 'destroy', 'validate_payment', 'reject_payment')
        if self.action in admin_actions:
            return [permissions.IsAdminUser()]
        return [permissions.AllowAny()]

    def get_serializer_class(self):
        return ContributionSerializer

    @action(
        detail=False,
        methods=['post'],
        url_path='student',
        parser_classes=[MultiPartParser, FormParser, JSONParser],
    )
    def create_student(self, request):
        serializer = StudentCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        try:
            contribution = serializer.save()
        except (OSError, IOError, PermissionError) as exc:
            return Response(
                {
                    'photo': [
                        'Impossible d\'enregistrer la photo sur le serveur. '
                        'Réessayez sans photo ou avec une image JPG plus légère.',
                    ],
                    'detail': str(exc) if settings.DEBUG else None,
                },
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )
        return Response(
            ContributionSerializer(contribution, context={'request': request}).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=False, methods=['post'], url_path='donor')
    def create_donor(self, request):
        serializer = DonorCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        contribution = serializer.save()
        return Response(
            ContributionSerializer(contribution).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=['post'], url_path='submit-donation')
    def submit_donation(self, request, pk=None):
        """Enregistre l'intention de don (sans paiement en ligne). Validation admin requise."""
        contribution = self.get_object()
        if contribution.contributor_type != 'donateur':
            return Response(
                {'detail': 'Réservé aux donateurs.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        serializer = PaymentRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        amount = serializer.validated_data['amount']
        phone = serializer.validated_data['phone']
        method = serializer.validated_data['method']
        min_amount = min(getattr(settings, 'MINIMUM_CONTRIBUTION_AMOUNT', 1000), 1000)
        if amount < min_amount:
            return Response(
                {'detail': f'Le montant minimum est de {min_amount} FCFA.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        contribution.amount = amount
        contribution.telephone = phone.strip()
        contribution.payment_method = method
        contribution.mode_paiement = 'en_ligne'
        contribution.payment_status = 'PENDING'
        contribution.amount_paid = 0
        contribution.save()
        return Response(ContributionSerializer(contribution).data)

    @action(detail=True, methods=['post'])
    def pay(self, request, pk=None):
        contribution = self.get_object()
        serializer = PaymentRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        amount = serializer.validated_data['amount']
        phone = serializer.validated_data['phone']
        method = serializer.validated_data['method']

        if contribution.contributor_type == 'etudiant':
            if amount != STUDENT_FEE:
                return Response(
                    {'detail': f"Le paiement étudiant doit être de {STUDENT_FEE} FCFA."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            min_amount = settings.MINIMUM_CONTRIBUTION_AMOUNT
            if amount < min_amount:
                return Response(
                    {'detail': f"Le montant minimum est de {min_amount} FCFA."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        client = TaramoneyClient()
        try:
            taramoney_response = client.initiate_payment(
                business_id=getattr(settings, 'TARAMONEY_BUSINESS_ID', ''),
                product_id=f"contrib-{contribution.id}",
                product_name=f"Paiement {contribution.get_contributor_type_display() or 'Contribution'}",
                product_price=amount,
                phone_number=phone,
                web_hook_url=getattr(settings, 'TARAMONEY_DEFAULT_WEBHOOK', ''),
                network='',
            )
        except TaramoneyError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_502_BAD_GATEWAY)

        if contribution.contributor_type == 'etudiant':
            if contribution.amount is None:
                contribution.amount = STUDENT_FEE
            PaymentInstallment.objects.create(
                contribution=contribution,
                installment_number=contribution.installments.count() + 1,
                amount=amount,
                mode_paiement='en_ligne',
                payment_method=method,
                campay_reference=taramoney_response.get('paymentId', ''),
                payment_status='PENDING',
            )
            contribution.mode_paiement = 'en_ligne'
            contribution.campay_reference = taramoney_response.get('paymentId', '')
            contribution.save(update_fields=['mode_paiement', 'campay_reference', 'updated_at'])

            contribution.recalculate_payment_status()
        else:
            contribution.amount = amount
            contribution.payment_method = method
            contribution.campay_reference = taramoney_response.get('paymentId', '')
            contribution.payment_status = 'PENDING'
            contribution.mode_paiement = 'en_ligne'
            contribution.save()

        return Response({
            'reference': contribution.campay_reference,
            'status': contribution.payment_status,
            'taramoney': taramoney_response,
        })

    @action(detail=True, methods=['post'], url_path='confirm-cash')
    def confirm_cash(self, request, pk=None):
        """Enregistre une demande de paiement en espèces (validation admin requise)."""
        contribution = self.get_object()
        if contribution.contributor_type != 'etudiant':
            return Response({'detail': 'Réservé aux inscriptions étudiantes.'}, status=status.HTTP_400_BAD_REQUEST)

        contribution.mode_paiement = 'especes'
        contribution.amount = STUDENT_FEE
        contribution.payment_status = 'PENDING'
        contribution.save(update_fields=['mode_paiement', 'amount', 'payment_status', 'updated_at'])
        return Response(ContributionSerializer(contribution).data)

    @action(
        detail=True,
        methods=['post'],
        url_path='upload-proof',
        parser_classes=[MultiPartParser, FormParser],
    )
    def upload_proof(self, request, pk=None):
        contribution = self.get_object()
        proof = request.FILES.get('payment_proof')
        if not proof:
            return Response({'detail': 'Fichier de preuve requis.'}, status=status.HTTP_400_BAD_REQUEST)

        contribution.payment_proof = proof
        contribution.mode_paiement = 'especes'
        contribution.amount = contribution.amount or STUDENT_FEE
        contribution.payment_status = 'PENDING'
        contribution.save()
        return Response(ContributionSerializer(contribution).data)

    @action(detail=True, methods=['get'], url_path='status')
    def payment_status_view(self, request, pk=None):
        contribution = self.get_object()
        if contribution.contributor_type == 'etudiant':
            pending_installments = contribution.installments.filter(payment_status='PENDING').exclude(campay_reference='')
            changed = False
            for installment in pending_installments:
                try:
                    live_status = campay.check_status(installment.campay_reference)
                    if live_status != installment.payment_status:
                        installment.payment_status = live_status
                        installment.save(update_fields=['payment_status'])
                        changed = True
                except campay.CampayError:
                    continue
            if changed:
                contribution.sync_amount_paid()
                contribution.recalculate_payment_status()
                contribution.save(update_fields=['amount_paid', 'payment_status', 'updated_at'])
        elif contribution.campay_reference and contribution.payment_status == 'PENDING':
            try:
                live_status = campay.check_status(contribution.campay_reference)
                if live_status != contribution.payment_status:
                    contribution.payment_status = live_status
                    contribution.save(update_fields=['payment_status'])
            except campay.CampayError:
                pass
        return Response({'status': contribution.payment_status})

    @action(detail=True, methods=['post'], url_path='validate-payment')
    def validate_payment(self, request, pk=None):
        """Valide manuellement un paiement (espèces, don en attente, etc.)."""
        contribution = self.get_object()
        if contribution.payment_status == 'SUCCESSFUL':
            return Response({'detail': 'Déjà validé.'}, status=status.HTTP_400_BAD_REQUEST)
        if not contribution.amount:
            contribution.amount = STUDENT_FEE if contribution.contributor_type == 'etudiant' else contribution.amount
        if contribution.contributor_type == 'donateur' and contribution.amount:
            contribution.amount_paid = contribution.amount
        elif contribution.contributor_type == 'etudiant' and not contribution.amount_paid:
            contribution.amount_paid = contribution.amount or STUDENT_FEE
        contribution.payment_status = 'SUCCESSFUL'
        contribution.save()
        return Response(ContributionSerializer(contribution).data)

    @action(detail=True, methods=['post'], url_path='reject-payment')
    def reject_payment(self, request, pk=None):
        contribution = self.get_object()
        contribution.payment_status = 'FAILED'
        contribution.save(update_fields=['payment_status', 'updated_at'])
        return Response(ContributionSerializer(contribution).data)

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAdminUser])
    def export_excel(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contributions"

        headers = [
            'Type', 'Nom', 'Prénom', 'Téléphone', 'Email', 'Méthode préférée',
            'École', 'Filière', 'Niveau académique', 'Ville de résidence',
            'Déjà participé', 'Ressortissant Est', 'Parle Maka’a', 'Taille T-shirt',
            'Allergies santé', 'Attentes campagne',
            'Contact urgence nom', 'Contact urgence lien', 'Contact urgence téléphone', 'Contact urgence ville',
            'Photo', 'Numéro carte étudiant', 'Mode paiement', 'Méthode',
            'Montant (FCFA)', 'Montant payé (FCFA)', 'Statut paiement', 'Référence Campay',
            'Preuve paiement', 'Date création', 'Date mise à jour',
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1B4D3E')

        for c in Contribution.objects.all().order_by('-created_at'):
            ws.append([
                c.get_contributor_type_display() if c.contributor_type else '',
                c.nom, c.prenom, c.telephone, c.email,
                c.get_methode_preferee_display(),
                c.ecole, c.filiere,
                c.get_niveau_academique_display() if c.niveau_academique else '',
                c.ville_residence, c.deja_participe_campagne, c.ressortissant_est,
                c.parle_makaa, c.taille_tshirt, c.allergies_sante, c.attentes_campagne,
                c.contact_urgence_nom, c.contact_urgence_lien,
                c.contact_urgence_telephone, c.contact_urgence_ville,
                c.photo.url if c.photo else '', c.numero_carte_etudiant,
                c.get_mode_paiement_display() if c.mode_paiement else '',
                c.get_methode_preferee_display(), c.amount or 0,
                c.amount_paid or 0, c.get_payment_status_display(), c.campay_reference,
                c.payment_proof.url if c.payment_proof else '',
                c.created_at.strftime('%d/%m/%Y %H:%M'),
                c.updated_at.strftime('%d/%m/%Y %H:%M'),
            ])

        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="contributions_bouano_doumaintang.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAdminUser])
    def export_csv(self, request):
        headers = [
            'Type', 'Nom', 'Prénom', 'Téléphone', 'Email', 'Méthode préférée',
            'École', 'Filière', 'Niveau académique', 'Ville de résidence',
            'Déjà participé', 'Ressortissant Est', 'Parle Maka’a', 'Taille T-shirt',
            'Allergies santé', 'Attentes campagne',
            'Contact urgence nom', 'Contact urgence lien', 'Contact urgence téléphone', 'Contact urgence ville',
            'Photo', 'Numéro carte étudiant', 'Mode paiement', 'Méthode',
            'Montant (FCFA)', 'Montant payé (FCFA)', 'Statut paiement', 'Référence Campay',
            'Preuve paiement', 'Date création', 'Date mise à jour',
        ]
        rows = [headers]
        for c in Contribution.objects.all().order_by('-created_at'):
            rows.append([
                c.get_contributor_type_display() if c.contributor_type else '',
                c.nom, c.prenom, c.telephone, c.email,
                c.get_methode_preferee_display(),
                c.ecole, c.filiere,
                c.get_niveau_academique_display() if c.niveau_academique else '',
                c.ville_residence, c.deja_participe_campagne, c.ressortissant_est,
                c.parle_makaa, c.taille_tshirt, c.allergies_sante, c.attentes_campagne,
                c.contact_urgence_nom, c.contact_urgence_lien,
                c.contact_urgence_telephone, c.contact_urgence_ville,
                c.photo.url if c.photo else '', c.numero_carte_etudiant,
                c.get_mode_paiement_display() if c.mode_paiement else '',
                c.get_methode_preferee_display(), c.amount or 0,
                c.amount_paid or 0, c.get_payment_status_display(), c.campay_reference,
                c.payment_proof.url if c.payment_proof else '',
                c.created_at.strftime('%d/%m/%Y %H:%M'),
                c.updated_at.strftime('%d/%m/%Y %H:%M'),
            ])
        import csv
        import io

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        for row in rows:
            writer.writerow(row)

        response = HttpResponse(buffer.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="contributions_bouano_doumaintang.csv"'
        return response

    @action(detail=True, methods=['get'], permission_classes=[permissions.IsAdminUser])
    def export_detail_excel(self, request, pk=None):
        """Export a single contribution's details as an Excel file."""
        c = self.get_object()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contribution"

        rows = [
            ('Champ', 'Valeur'),
            ('Type', c.get_contributor_type_display() if c.contributor_type else ''),
            ('Nom', c.nom),
            ('Prénom', c.prenom),
            ('Téléphone', c.telephone),
            ('Email', c.email),
            ('Méthode préférée', c.get_methode_preferee_display()),
            ('École', c.ecole),
            ('Filière', c.filiere),
            ('Niveau académique', c.get_niveau_academique_display() if c.niveau_academique else ''),
            ('Ville de résidence', c.ville_residence),
            ('Déjà participé', c.deja_participe_campagne),
            ('Ressortissant Est', c.ressortissant_est),
            ("Parle Maka'a", c.parle_makaa),
            ('Taille T-shirt', c.taille_tshirt),
            ('Allergies santé', c.allergies_sante),
            ('Attentes campagne', c.attentes_campagne),
            ('Contact urgence nom', c.contact_urgence_nom),
            ('Contact urgence téléphone', c.contact_urgence_telephone),
            ('Contact urgence ville', c.contact_urgence_ville),
            ('Contact urgence lien', c.contact_urgence_lien),
            ('Photo URL', c.photo.url if c.photo else ''),
            ('Numéro carte étudiant', c.numero_carte_etudiant),
            ('Mode paiement', c.get_mode_paiement_display() if c.mode_paiement else ''),
            ('Méthode', c.get_methode_preferee_display()),
            ('Montant (FCFA)', c.amount or 0),
            ('Montant payé (FCFA)', c.amount_paid or 0),
            ('Statut paiement', c.get_payment_status_display()),
            ('Référence Campay', c.campay_reference),
            ('Preuve paiement URL', c.payment_proof.url if c.payment_proof else ''),
            ('Date création', c.created_at.strftime('%d/%m/%Y %H:%M')),
            ('Date mise à jour', c.updated_at.strftime('%d/%m/%Y %H:%M')),
        ]

        for row in rows:
            ws.append(row)

        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1B4D3E')

        # Auto width
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"contribution_{c.id}_details.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def api_health(request):
    return Response({'status': 'ok', 'live': True})


class StartTaraPayment(APIView):
    """Start a payment via Taramoney (TaraMoney) by forwarding the expected payload.

    Expected JSON body:
      - productId (string)
      - productName (string)
      - productPrice (int)
      - phoneNumber (string, with country code)
      - webHookUrl (string) optional — falls back to settings.TARAMONEY_DEFAULT_WEBHOOK
      - network (string) optional
    """
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        data = request.data
        product_id = data.get('productId')
        product_name = data.get('productName')
        product_price = data.get('productPrice')
        phone_number = data.get('phoneNumber')
        web_hook = data.get('webHookUrl') or getattr(settings, 'TARAMONEY_DEFAULT_WEBHOOK', '')
        network = data.get('network', '')

        if not all([product_id, product_name, product_price, phone_number]):
            return Response({'detail': 'Champs requis manquants.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            product_price = int(product_price)
        except (TypeError, ValueError):
            return Response({'detail': 'productPrice doit être un entier.'}, status=status.HTTP_400_BAD_REQUEST)

        client = TaramoneyClient()
        try:
            resp = client.initiate_payment(
                business_id=getattr(settings, 'TARAMONEY_BUSINESS_ID', ''),
                product_id=product_id,
                product_name=product_name,
                product_price=product_price,
                phone_number=phone_number,
                web_hook_url=web_hook,
                network=network,
            )
        except TaramoneyError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_502_BAD_GATEWAY)

        return Response(resp)


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def campay_webhook(request):
    reference = request.data.get('reference')
    external_status = request.data.get('status')
    if not reference or not external_status:
        return Response({'detail': 'Payload invalide'}, status=status.HTTP_400_BAD_REQUEST)

    installment = PaymentInstallment.objects.filter(campay_reference=reference).first()
    if installment:
        installment.payment_status = external_status
        installment.save(update_fields=['payment_status'])
        contribution = installment.contribution
        contribution.sync_amount_paid()
        contribution.recalculate_payment_status()
        contribution.save(update_fields=['amount_paid', 'payment_status', 'updated_at'])
        return Response({'ok': True})

    try:
        contribution = Contribution.objects.get(campay_reference=reference)
    except Contribution.DoesNotExist:
        return Response({'detail': 'Contribution introuvable'}, status=status.HTTP_404_NOT_FOUND)

    contribution.payment_status = external_status
    contribution.save(update_fields=['payment_status'])
    return Response({'ok': True})


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def taramoney_webhook(request):
    payload = request.data if isinstance(request.data, dict) else {}
    payment_id = payload.get('paymentId') or payload.get('payment_id')
    business_id = payload.get('businessId') or payload.get('business_id')
    collection_id = payload.get('collectionId') or payload.get('collection_id')
    phone = payload.get('phoneNumber') or payload.get('phone_number')
    status_val = payload.get('status')

    if not payment_id and not collection_id:
        return Response({'detail': 'Payload invalide'}, status=status.HTTP_400_BAD_REQUEST)

    p = Payment.objects.create(
        business_id=business_id or '',
        payment_id=payment_id or '',
        collection_id=collection_id or '',
        phone_number=phone or '',
        status=status_val or '',
        raw_payload=payload,
    )

    # Try linking to an installment by reference or by phone
    linked = False
    if payment_id:
        inst = PaymentInstallment.objects.filter(campay_reference=payment_id).first()
        if inst:
            p.installment = inst
            p.contribution = inst.contribution
            linked = True
    if not linked and phone:
        contrib = Contribution.objects.filter(telephone__endswith=phone).first()
        if contrib:
            p.contribution = contrib
            linked = True

    p.save()

    # Update linked records
    if p.installment:
        p.installment.payment_status = status_val or p.installment.payment_status
        p.installment.save(update_fields=['payment_status'])
        contrib = p.installment.contribution
        contrib.sync_amount_paid()
        contrib.recalculate_payment_status()
        contrib.save(update_fields=['amount_paid', 'payment_status', 'updated_at'])
    elif p.contribution:
        p.contribution.payment_status = status_val or p.contribution.payment_status
        p.contribution.save(update_fields=['payment_status', 'updated_at'])

    return Response({'ok': True})


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def campaign_stats(request):
    paid = Contribution.objects.filter(payment_status='SUCCESSFUL')
    total_collected = paid.aggregate(total=models.Sum('amount_paid'))['total'] or 0
    students_count = Contribution.objects.filter(
        contributor_type='etudiant',
        payment_status__in=['SUCCESSFUL', 'PENDING', 'PARTIAL'],
    ).count()
    return Response({
        'total_collected': total_collected,
        'contributors_count': paid.count(),
        'students_registered': students_count,
        'goal': settings.FUNDRAISING_GOAL,
    })


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def registered_students(request):
    """Liste publique des étudiants inscrits (sans données sensibles)."""
    # Include students who are either SUCCESSFUL, or who paid in cash and uploaded a proof
    qs = Contribution.objects.filter(contributor_type='etudiant').filter(
        models.Q(payment_status='SUCCESSFUL')
        | (models.Q(mode_paiement='especes') & ~models.Q(payment_proof=''))
    ).exclude(payment_status='FAILED')
    serializer = RegisteredStudentSerializer(qs[:100], many=True)
    return Response(serializer.data)
