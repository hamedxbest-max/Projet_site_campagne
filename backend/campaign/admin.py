from django.contrib import admin
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill

from .models import Contribution, PaymentInstallment


@admin.action(description="Exporter la sélection en fichier Excel (.xlsx)")
def export_to_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contributions"
    headers = [
        'Type', 'Nom', 'Prénom', 'Téléphone', 'Email', 'Méthode préférée',
        'École', 'Filière', 'Niveau académique', 'Ville de résidence',
        'Déjà participé', 'Ressortissant Est', 'Parle Maka’a', 'Taille T-shirt',
        'Allergies santé', 'Attentes campagne',
        'Contact urgence nom', 'Contact urgence lien', 'Contact urgence téléphone', 'Contact urgence ville',
        'Photo', 'Numéro carte étudiant', 'Mode paiement', 'Méthode',
        'Montant (FCFA)', 'Montant payé (FCFA)', 'Statut paiement', 'Référence Campay',
        'Preuve paiement', 'Date création', 'Date mise à jour',
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1B4D3E')

    for c in queryset.order_by('-created_at'):
        ws.append([
            c.get_contributor_type_display() if c.contributor_type else '',
            c.nom, c.prenom, c.telephone, c.email,
            c.get_methode_preferee_display(),
            c.ecole, c.filiere,
            c.get_niveau_academique_display() if c.niveau_academatique else '',
            c.ville_residence, c.deja_participe_campagne, c.ressortissant_est,
            c.parle_makaa, c.taille_tshirt, c.allergies_sante, c.attentes_campagne,
            c.contact_urgence_nom, c.contact_urgence_lien,
            c.contact_urgence_telephone, c.contact_urgence_ville,
            c.photo.url if c.photo else '', c.numero_carte_etudiant,
            c.get_mode_paiement_display() if c.mode_paiement else '',
            c.get_methode_preferee_display(), c.amount or 0,
            c.amount_paid or 0, c.get_payment_status_display(), c.campay_reference,
            c.payment_proof.url if c.payment_proof else '',
            c.created_at.strftime('%d/%m/%Y %H:%M'),
            c.updated_at.strftime('%d/%m/%Y %H:%M'),
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="contributions_bouano_doumaintang.xlsx"'
    wb.save(response)
    return response


class PaymentInstallmentInline(admin.TabularInline):
    model = PaymentInstallment
    fields = ('installment_number', 'amount', 'mode_paiement', 'payment_method', 'payment_status', 'campay_reference', 'created_at')
    readonly_fields = ('installment_number', 'amount', 'mode_paiement', 'payment_method', 'payment_status', 'campay_reference', 'created_at')
    can_delete = False
    extra = 0

@admin.register(Contribution)
class ContributionAdmin(admin.ModelAdmin):
    list_display = (
        'display_name', 'contributor_type', 'ecole', 'niveau_academique', 'telephone',
        'email', 'amount', 'amount_paid', 'mode_paiement', 'payment_status', 'created_at',
    )
    list_filter = (
        'payment_status', 'contributor_type', 'mode_paiement', 'payment_method', 'niveau_academique',
    )
    search_fields = (
        'nom', 'prenom', 'telephone', 'email', 'campay_reference', 'numero_carte_etudiant',
        'ecole', 'filiere', 'ville_residence', 'contact_urgence_nom', 'contact_urgence_telephone',
    )
    actions = [export_to_excel]
    readonly_fields = ('created_at', 'updated_at')
    inlines = [PaymentInstallmentInline]
    fieldsets = (
        ('Informations personnelles', {
            'fields': (
                'contributor_type', 'nom', 'prenom', 'age', 'sexe', 'telephone', 'email',
                'methode_preferee', 'photo', 'numero_carte_etudiant',
            )
        }),
        ('Détails académiques', {
            'fields': (
                'niveau_academique', 'ecole', 'filiere', 'ville_residence',
            )
        }),
        ('Préférences de campagne', {
            'fields': (
                'deja_participe_campagne', 'ressortissant_est', 'parle_makaa', 'taille_tshirt',
                'allergies_sante', 'attentes_campagne',
            )
        }),
        ('Contact d\'urgence', {
            'fields': (
                'contact_urgence_nom', 'contact_urgence_lien',
                'contact_urgence_telephone', 'contact_urgence_ville',
            )
        }),
        ('Paiement', {
            'fields': (
                'mode_paiement', 'payment_method', 'amount', 'amount_paid',
                'payment_status', 'campay_reference', 'payment_proof',
            )
        }),
        ('Dates', {
            'fields': ('created_at', 'updated_at'),
        }),
    )

    @admin.display(description='Nom complet')
    def display_name(self, obj):
        return obj.display_name
